"""Recalculate and verify one XLSX workbook with LibreOffice.

This is a clean-room Code implementation using the public XLSX package format,
openpyxl, and LibreOffice's command-line interface.  It does not search outside
its own resource directory or the explicitly supplied workbook path.
"""

from __future__ import annotations

import argparse
import html
import json
import os
import re
import shutil
import stat
import sys
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path

from office.soffice import SofficeError, run_soffice_conversion


SCHEMA = "code-xlsx-recalc/v1"
DEFAULT_TIMEOUT_SECONDS = 30
MIN_TIMEOUT_SECONDS = 1
MAX_TIMEOUT_SECONDS = 300
MAX_WORKBOOK_BYTES = 256 * 1024 * 1024
MAX_ERROR_LOCATIONS_PER_TYPE = 100
MAX_FORMULA_XML_BYTES = 64 * 1024 * 1024
FORMULA_ERRORS = {
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#GETTING_DATA",
    "#SPILL!", "#CALC!", "#FIELD!", "#BLOCKED!", "#UNKNOWN!", "#CONNECT!", "#BUSY!",
}
_FORMULA_XML = re.compile(rb"<f(?:\s[^>]*)?>(.*?)</f>", re.IGNORECASE | re.DOTALL)
_EXTERNAL_FORMULA = re.compile(
    r"(?i)(?:\[[^\]]+\]|\b(?:WEBSERVICE|DDE|RTD)\s*\(|(?:https?|ftp|file)://|\\\\)",
)

_PUBLIC_ERRORS = {
    "invalid_arguments": "Usage: recalc.py WORKBOOK.xlsx [timeout_seconds] [--force].",
    "invalid_timeout": "timeout_seconds must be an integer between 1 and 300.",
    "workbook_not_found": "The workbook does not exist or is not a regular file.",
    "workbook_type_unsupported": "Only .xlsx workbooks can be recalculated.",
    "workbook_too_large": "The workbook is too large to recalculate safely.",
    "workbook_invalid": "The workbook is not a readable XLSX package.",
    "external_links_detected": "External workbook links were detected. Recalculation is blocked unless --force explicitly accepts link loss.",
    "openpyxl_missing": "openpyxl is required for XLSX formula verification.",
    "soffice_not_found": "LibreOffice is required for formula recalculation but was not found.",
    "soffice_invalid": "The selected LibreOffice executable is not a supported installation.",
    "timeout": "LibreOffice recalculation timed out.",
    "process_failed": "LibreOffice could not recalculate the workbook.",
    "output_missing": "LibreOffice did not produce the expected workbook.",
    "result_invalid": "The recalculated workbook could not be verified.",
    "replace_failed": "The recalculated workbook could not be published in place.",
    "cancelled": "Workbook recalculation was cancelled.",
}


class RecalcError(RuntimeError):
    def __init__(self, error_code):
        self.error_code = str(error_code or "result_invalid")
        super().__init__(_PUBLIC_ERRORS.get(self.error_code, _PUBLIC_ERRORS["result_invalid"]))


def _is_link_or_reparse(path):
    try:
        metadata = os.lstat(path)
    except OSError:
        return False
    if stat.S_ISLNK(metadata.st_mode):
        return True
    attributes = int(getattr(metadata, "st_file_attributes", 0) or 0)
    reparse_flag = int(getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400) or 0x400)
    return bool(attributes & reparse_flag)


def _bounded_timeout(value):
    try:
        timeout = int(str(value or DEFAULT_TIMEOUT_SECONDS))
    except (TypeError, ValueError):
        raise RecalcError("invalid_timeout")
    if not MIN_TIMEOUT_SECONDS <= timeout <= MAX_TIMEOUT_SECONDS:
        raise RecalcError("invalid_timeout")
    return timeout


def _validate_workbook(path):
    path = Path(path)
    if not path.is_file() or _is_link_or_reparse(path):
        raise RecalcError("workbook_not_found")
    if path.suffix.lower() != ".xlsx":
        raise RecalcError("workbook_type_unsupported")
    try:
        if path.stat().st_size > MAX_WORKBOOK_BYTES:
            raise RecalcError("workbook_too_large")
    except OSError:
        raise RecalcError("workbook_not_found")
    if not zipfile.is_zipfile(path):
        raise RecalcError("workbook_invalid")
    return path.resolve(strict=True)


def detect_external_links(path):
    """Return bounded external-link package markers without resolving links."""
    markers = []
    try:
        with zipfile.ZipFile(path, "r") as archive:
            names = archive.namelist()
            for name in names:
                normalized = name.replace("\\", "/")
                if (
                    normalized.startswith("xl/externalLinks/")
                    or normalized.startswith("xl/queryTables/")
                    or normalized == "xl/connections.xml"
                ):
                    markers.append(normalized)
                    if len(markers) >= 20:
                        break
            if len(markers) < 20 and "xl/_rels/workbook.xml.rels" in names:
                relationships = archive.read("xl/_rels/workbook.xml.rels")
                if b'TargetMode="External"' in relationships or b"TargetMode='External'" in relationships:
                    markers.append("xl/_rels/workbook.xml.rels")
            if len(markers) < 20:
                for name in names:
                    normalized = name.replace("\\", "/")
                    if not normalized.startswith("xl/worksheets/") or not normalized.endswith(".xml"):
                        continue
                    info = archive.getinfo(name)
                    if info.file_size > MAX_FORMULA_XML_BYTES:
                        raise RecalcError("workbook_invalid")
                    payload = archive.read(name)
                    if any(
                        _EXTERNAL_FORMULA.search(html.unescape(match.decode("utf-8", "replace")))
                        for match in _FORMULA_XML.findall(payload)
                    ):
                        markers.append(f"formula:{normalized}")
                        if len(markers) >= 20:
                            break
    except (OSError, zipfile.BadZipFile, KeyError):
        raise RecalcError("workbook_invalid")
    return markers


def _load_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RecalcError("openpyxl_missing")
    return load_workbook


def _scan_result(path):
    load_workbook = _load_openpyxl()
    formula_count = 0
    errors = defaultdict(list)
    error_counts = defaultdict(int)
    total_errors = 0
    try:
        formula_book = load_workbook(path, data_only=False, read_only=True)
        value_book = load_workbook(path, data_only=True, read_only=True)
        try:
            for formula_sheet in formula_book.worksheets:
                value_sheet = value_book[formula_sheet.title]
                for formula_row, value_row in zip(formula_sheet.iter_rows(), value_sheet.iter_rows()):
                    for formula_cell, value_cell in zip(formula_row, value_row):
                        if formula_cell.data_type == "f" or (
                            isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
                        ):
                            formula_count += 1
                        value = value_cell.value
                        if value_cell.data_type == "e" or value in FORMULA_ERRORS:
                            error_name = str(value or "#UNKNOWN!")
                            total_errors += 1
                            error_counts[error_name] += 1
                            if len(errors[error_name]) < MAX_ERROR_LOCATIONS_PER_TYPE:
                                errors[error_name].append(f"{formula_sheet.title}!{formula_cell.coordinate}")
        finally:
            formula_book.close()
            value_book.close()
    except RecalcError:
        raise
    except Exception:
        raise RecalcError("result_invalid")
    summary = {
        name: {
            "locations": locations,
            "locations_truncated": max(0, error_counts[name] - len(locations)),
        }
        for name, locations in errors.items()
    }
    return formula_count, total_errors, summary


def recalculate_with_libreoffice(source, destination, *, timeout_seconds, soffice_path="", cancel_event=None):
    with tempfile.TemporaryDirectory(prefix="code-xlsx-recalc-") as temp:
        root = Path(temp)
        input_dir = root / "input"
        output_dir = root / "output"
        profile_dir = root / "profile"
        input_dir.mkdir()
        input_copy = input_dir / source.name
        shutil.copy2(source, input_copy)
        output = run_soffice_conversion(
            input_copy,
            output_dir,
            profile_dir,
            timeout_seconds=timeout_seconds,
            executable=soffice_path,
            cancel_event=cancel_event,
        )
        shutil.copy2(output, destination)


def _publish_staged_workbook(staged, workbook):
    publish_path = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            prefix=f".{workbook.name}.",
            suffix=".recalc.tmp",
            dir=workbook.parent,
            delete=False,
        ) as stream:
            publish_path = Path(stream.name)
            with Path(staged).open("rb") as source:
                shutil.copyfileobj(source, stream, length=128 * 1024)
            stream.flush()
            os.fsync(stream.fileno())
        try:
            shutil.copymode(workbook, publish_path)
        except OSError:
            pass
        os.replace(publish_path, workbook)
        publish_path = None
    except RecalcError:
        raise
    except OSError:
        raise RecalcError("replace_failed")
    finally:
        if publish_path is not None:
            try:
                publish_path.unlink(missing_ok=True)
            except OSError:
                pass


def recalculate_workbook(path, timeout_seconds=DEFAULT_TIMEOUT_SECONDS, *, force=False, soffice_path="", cancel_event=None):
    workbook = _validate_workbook(path)
    timeout = _bounded_timeout(timeout_seconds)
    if detect_external_links(workbook) and not force:
        raise RecalcError("external_links_detected")
    try:
        with tempfile.TemporaryDirectory(prefix="code-xlsx-publish-") as temp:
            staged = Path(temp) / workbook.name
            recalculate_with_libreoffice(
                workbook,
                staged,
                timeout_seconds=timeout,
                soffice_path=soffice_path,
                cancel_event=cancel_event,
            )
            staged = _validate_workbook(staged)
            formulas, total_errors, summary = _scan_result(staged)
            _publish_staged_workbook(staged, workbook)
    except SofficeError as exc:
        raise RecalcError(exc.error_code)
    return {
        "schema": SCHEMA,
        "status": "errors_found" if total_errors else "success",
        "total_formulas": formulas,
        "total_errors": total_errors,
        "error_summary": summary,
    }


def _write_json(payload):
    sys.stdout.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")


def main(argv=None):
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("workbook", nargs="?")
    parser.add_argument("timeout", nargs="?", default=str(DEFAULT_TIMEOUT_SECONDS))
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--soffice", default="")
    try:
        arguments = parser.parse_args(argv)
        if not arguments.workbook:
            raise RecalcError("invalid_arguments")
        timeout = _bounded_timeout(arguments.timeout)
        result = recalculate_workbook(
            arguments.workbook,
            timeout,
            force=arguments.force,
            soffice_path=arguments.soffice,
        )
        _write_json(result)
        return 0
    except SystemExit:
        _write_json({"schema": SCHEMA, "errorCode": "invalid_arguments", "error": _PUBLIC_ERRORS["invalid_arguments"]})
        return 2
    except KeyboardInterrupt:
        error = RecalcError("cancelled")
        _write_json({"schema": SCHEMA, "errorCode": error.error_code, "error": str(error)})
        return 130
    except RecalcError as exc:
        _write_json({"schema": SCHEMA, "errorCode": exc.error_code, "error": str(exc)})
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
